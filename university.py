# ==========================================================
# university.py
# Equivalente a university.R sin pandas
# Requiere:
# pip install openpyxl
# ==========================================================

import sqlite3
from openpyxl import load_workbook

# ------------------------------------------
# Ruta del archivo Excel
# ------------------------------------------

archivo = "/Users/agodoy/notps/Ecuola/analisis_ciencia_de_datos/sql_taller/university.xlsx"

# ------------------------------------------
# Función similar a read_excel()
# ------------------------------------------


def read_sheet(workbook, sheet_name):

    hoja = workbook[sheet_name]

    filas = list(hoja.iter_rows(values_only=True))

    encabezados = filas[0]

    datos = []

    for fila in filas[1:]:
        registro = dict(zip(encabezados, fila))
        datos.append(registro)

    return datos


# ------------------------------------------
# Leer todas las hojas
# ------------------------------------------

wb = load_workbook(archivo, data_only=True)

instructor = read_sheet(wb, "instructor")
course = read_sheet(wb, "course")
teaches = read_sheet(wb, "teaches")
section = read_sheet(wb, "section")
department = read_sheet(wb, "department")
prereq = read_sheet(wb, "prereq")
classroom = read_sheet(wb, "classroom")
student = read_sheet(wb, "student")
takes = read_sheet(wb, "takes")
advisor = read_sheet(wb, "advisor")
time_slot = read_sheet(wb, "time_slot")


# ------------------------------------------
# Crear conexión SQLite
# ------------------------------------------

con = sqlite3.connect("university_profesor.db")

cursor = con.cursor()

# Activar claves foráneas
cursor.execute("PRAGMA foreign_keys=ON;")


# ------------------------------------------
# Crear tablas
# ------------------------------------------

cursor.execute("""

CREATE TABLE IF NOT EXISTS department (
    dept_name TEXT PRIMARY KEY,
    building TEXT,
    budget REAL

)

""")

cursor.execute("""

CREATE TABLE IF NOT EXISTS instructor (

    ID INTEGER PRIMARY KEY,
    name TEXT,
    dept_name TEXT,
    salary REAL,

    FOREIGN KEY(dept_name)
    REFERENCES department(dept_name)

)

""")


cursor.execute("""

CREATE TABLE IF NOT EXISTS student (

    ID TEXT PRIMARY KEY,
    name TEXT,
    dept_name TEXT,
    tot_cred TEXT,
    advisor_ID TEXT,

    FOREIGN KEY(dept_name)
    REFERENCES department(dept_name)

)

""")


cursor.execute("""

CREATE TABLE IF NOT EXISTS advisor (

    s_ID TEXT,
    i_ID INTEGER,

    PRIMARY KEY(s_ID,i_ID),

    FOREIGN KEY(s_ID)
    REFERENCES student(ID),

    FOREIGN KEY(i_ID)
    REFERENCES instructor(ID)

)

""")


cursor.execute("""

CREATE TABLE IF NOT EXISTS course(

    course_id TEXT PRIMARY KEY,
    title TEXT,
    dept_name TEXT,
    credits INTEGER,

    FOREIGN KEY(dept_name)
    REFERENCES department(dept_name)

)

""")


cursor.execute("""

CREATE TABLE IF NOT EXISTS classroom(

    building TEXT,
    room_number TEXT,
    capacity INTEGER,

    PRIMARY KEY(building,room_number)

)

""")


cursor.execute("""

CREATE TABLE IF NOT EXISTS time_slot(

    time_slot_id TEXT,
    day TEXT,
    start_time TEXT,
    end_time TEXT,

    PRIMARY KEY(time_slot_id,day)

)

""")


cursor.execute("""

CREATE TABLE IF NOT EXISTS section(

    course_id TEXT,
    sec_id INTEGER,
    semester TEXT,
    year INTEGER,
    building TEXT,
    room_number TEXT,

    PRIMARY KEY(
        course_id,
        sec_id,
        semester,
        year
    ),

    FOREIGN KEY(course_id)
    REFERENCES course(course_id)


)

""")
#  FOREIGN KEY(building,room_number)
#  REFERENCES classroom(
#      building,
#      room_number


cursor.execute("""

CREATE TABLE IF NOT EXISTS prereq(

    course_id TEXT,
    prereq_id TEXT,

    PRIMARY KEY(
        course_id,
        prereq_id
    ),

    FOREIGN KEY(course_id)
    REFERENCES course(course_id),

    FOREIGN KEY(prereq_id)
    REFERENCES course(course_id)

)

""")


cursor.execute("""

CREATE TABLE IF NOT EXISTS teaches(

    ID INTEGER,
    course_id TEXT,
    sec_id INTEGER,
    semester TEXT,
    year INTEGER,

    PRIMARY KEY(
        ID,
        course_id,
        sec_id,
        semester,
        year
    ),

    FOREIGN KEY(ID)
    REFERENCES instructor(ID),

    FOREIGN KEY(course_id)
    REFERENCES course(course_id)

)

""")


cursor.execute("""

CREATE TABLE IF NOT EXISTS takes(

    ID TEXT,
    course_id TEXT,
    sec_id INTEGER,
    semester TEXT,
    year INTEGER,
    grade TEXT,

    PRIMARY KEY(
        ID,
        course_id,
        sec_id,
        semester,
        year
    ),

    FOREIGN KEY(ID)
    REFERENCES student(ID),

    FOREIGN KEY(course_id)
    REFERENCES course(course_id)

)

""")


# ------------------------------------------
# Función equivalente a dbWriteTable()
# ------------------------------------------


def insertar(tabla, datos):

    if len(datos) == 0:
        return

    columnas = list(datos[0].keys())

    campos = ",".join(columnas)

    placeholders = ",".join(["?"] * len(columnas))

    sql = f"""
    INSERT INTO {tabla}
    ({campos})
    VALUES ({placeholders})
    """

    valores = []

    for fila in datos:
        valores.append(tuple(fila[c] for c in columnas))

    cursor.executemany(sql, valores)


# ------------------------------------------
# Insertar datos
# ------------------------------------------

insertar("department", department)
insertar("instructor", instructor)
insertar("student", student)
insertar("advisor", advisor)
insertar("course", course)
insertar("classroom", classroom)
insertar("time_slot", time_slot)
insertar("section", section)
insertar("prereq", prereq)
insertar("teaches", teaches)
insertar("takes", takes)


# Guardar cambios
con.commit()

print("\nBase de datos SQLite 'university_python.db' creada correctamente.\n")

# Cerrar conexión
con.close()
